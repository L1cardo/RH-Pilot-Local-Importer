import logging
import re
from eventmanager import Evt
from data_import import DataImporter
from Database import ProgramMethod
from io import BytesIO
from openpyxl import load_workbook
from collections import defaultdict

logger = logging.getLogger(__name__)


def import_pilots(importer_class, rhapi, data, args):
    if not data:
        return False

    try:
        excel_file = BytesIO(data)
        workbook = load_workbook(filename=excel_file)
        sheet = workbook.active
    except Exception as ex:
        logger.error("Unable to import file: {}".format(str(ex)))
        return False

    heats = defaultdict(list)

    logger.info("Importing Pilots...")
    for index, row in enumerate(sheet.iter_rows(values_only=True)):
        if index == 0:
            class_name = row[0]
        elif index != 1:
            pilot_heat = row[0]
            pilot_name = row[1]
            pilot_callsign = row[1]
            pilot_frequency = row[2]
            pilot_color = get_pilot_color(row[3])
            if pilot_heat and pilot_name and pilot_callsign:
                pilot = {
                    "heat": pilot_heat,
                    "name": pilot_name,
                    "callsign": pilot_callsign,
                    "frequency": pilot_frequency,
                    "color": pilot_color,
                }

            existing_pilot = check_existing_pilot(rhapi, pilot)
            if not existing_pilot:
                rhapi.db.pilot_add(
                    name=pilot["name"], callsign=pilot["callsign"], color=pilot["color"]
                )
                logger.info(f"Pilot added: {pilot}")
            else:
                logger.info(f"Pilot alredy exists: {pilot}")

            pilot_id = get_pilot_id(rhapi, pilot)
            heats[pilot_heat].append(pilot_id)

    # rhapi.ui.broadcast_pilots()
    logger.info("Import Pilots complete")

    # Generate heats
    import_heats(rhapi, class_name, heats)

    return True


def import_heats(rhapi, class_name, heats):
    existing_class = check_existing_class(rhapi, class_name)

    if not existing_class:
        race_class = rhapi.db.raceclass_add(name=class_name)
        logger.info(f"Added Race Class: {class_name}")

        logger.info("Importing Heats...")
        # Create heats and associate them with the race class and format
        for heat_name, pilot_ids in heats.items():
            new_heat = rhapi.db.heat_add(name=heat_name)

            # Associate the heat with the race class and format
            rhapi.db.heat_alter(new_heat.id, raceclass=race_class.id)
            # Get all slots for this heat
            slots = rhapi.db.slots_by_heat(new_heat.id)

            for index, pilot_id in enumerate(pilot_ids):
                slot_id = slots[index].id
                rhapi.db.slot_alter(
                    slot_id, method=ProgramMethod.ASSIGN, pilot=pilot_id
                )
        rhapi.ui.message_notify(
            rhapi.__("Race Class, Pilots and Heats imported successfully")
        )
        logger.info("Race Class, Pilots and Heats imported successfully")

        # rhapi.ui.broadcast_raceclasses()
        # rhapi.ui.broadcast_heats()

    else:
        rhapi.ui.message_alert(
            rhapi.__("Race Class name exists, please change to another name")
        )
        logger.warning("Race Class name exists, please change to another name")


def check_existing_pilot(rhapi, pilot):
    existing_pilots = rhapi.db.pilots
    existing = False
    for existing_pilot in existing_pilots:
        if (
            existing_pilot.name == pilot["name"]
            and existing_pilot.callsign == pilot["callsign"]
        ):
            existing = True
    return existing


def check_existing_class(rhapi, class_name):
    existing_classes = rhapi.db.raceclasses
    existing = False
    for existing_class in existing_classes:
        if existing_class.name == class_name:
            existing = True
    return existing


def get_pilot_id(rhapi, pilot):
    db_pilots = rhapi.db.pilots
    for db_pilot in db_pilots:
        if db_pilot.name == pilot["name"] and db_pilot.callsign == pilot["callsign"]:
            return db_pilot.id
    return None


def get_pilot_color(color):
    pilot_color = "#FFFFFF"

    if is_hex_color(color):
        pilot_color = color
    else:
        if color in ("Blue", "蓝"):
            pilot_color = "#0022ff"
        elif color in ("Orange", "橙"):
            pilot_color = "#ff5500"
        elif color in ("Green", "绿"):
            pilot_color = "#00ff22"
        elif color in ("Red", "红"):
            pilot_color = "#ff0055"
        elif color in ("Yellow", "黄"):
            pilot_color = "#ddff00"
        elif color in ("Purple", "紫"):
            pilot_color = "#7700ff"
        elif color in ("Cyan", "青"):
            pilot_color = "#00ffdd"
        elif color in ("Grey", "灰"):
            pilot_color = "#aaaaaa"

    return pilot_color


def is_hex_color(color):
    if color is None or not isinstance(color, str):
        return False
    if color.startswith("#"):
        color = color[1:]
        if len(color) not in (6, 8):
            return False
    elif color.startswith("0x"):
        color = color[2:]
        if len(color) not in (6, 8):
            return False
    else:
        if len(color) not in (6, 8):
            return False
    hex_pattern = r"^[0-9A-Fa-f]+$"
    return bool(re.match(hex_pattern, color))


def register_handlers(args):
    args["register_fn"](
        DataImporter(
            "Import Pilots and Heats",
            import_pilots,
        )
    )


def initialize(rhapi):
    rhapi.events.on(Evt.DATA_IMPORT_INITIALIZE, register_handlers)
